from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def export_to_xlsx(extracted_data: dict, mappings: list[dict], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = [m["column_name"] for m in mappings]

    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 20

    # Build a row-per-entity output: gather all rows first
    rows: list[dict[str, object]] = []
    per_mapping_rows: dict[int, list] = {}

    for m_idx, mapping in enumerate(mappings):
        layer = mapping.get("layer")
        entity_type = mapping.get("entity_type")
        field = mapping.get("field", "text")

        if layer not in extracted_data.get("layers", {}):
            continue

        entities = [
            e for e in extracted_data["layers"][layer]["entities"]
            if e["type"] == entity_type
        ]
        per_mapping_rows[m_idx] = [(_get_field(e, field), e) for e in entities]

    # Determine max rows needed
    max_rows = max((len(v) for v in per_mapping_rows.values()), default=0)

    for row_idx in range(max_rows):
        excel_row = row_idx + 2
        for m_idx, mapping in enumerate(mappings):
            col_idx = m_idx + 1
            col_rows = per_mapping_rows.get(m_idx, [])
            if row_idx < len(col_rows):
                value, _ = col_rows[row_idx]
                if value is not None:
                    ws.cell(row=excel_row, column=col_idx, value=value)

    # Auto-width columns
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(output_path)


def _get_field(entity: dict, field: str):
    if field == "text":
        return entity.get("text")
    if field == "measurement":
        return entity.get("measurement")
    if field == "block_name":
        return entity.get("block_name")
    if field == "radius":
        return entity.get("radius")
    if field == "area":
        points = entity.get("points", [])
        if entity.get("is_closed") and len(points) >= 3:
            return round(_shoelace_area(points), 4)
        return None
    if field.startswith("attr:"):
        tag = field.split(":", 1)[1]
        return entity.get("attributes", {}).get(tag)
    return None


def _shoelace_area(points: list[list[float]]) -> float:
    n = len(points)
    area = 0.0
    for i in range(n):
        j = (i + 1) % n
        area += points[i][0] * points[j][1]
        area -= points[j][0] * points[i][1]
    return abs(area) / 2
