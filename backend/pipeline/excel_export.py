"""
Export a rebar schedule to .xlsx in the Romanian "Extras de armătură" format.

Columns (exact order):
  Marca | Ø [mm] | Oțel | Buc. | Lung. [m] | Lung./Ø [m] | Masa Ø/m [kg/m] | Masa/Ø [kg] | Masa totală [kg]

Layout:
  • Optional header block (project name, beneficiary, etc.)
  • Column header row
  • Data rows, grouped by diameter (sorted ascending)
  • Subtotal row per diameter group
  • Grand total row at the bottom
"""

from __future__ import annotations

import logging
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Styling helpers ────────────────────────────────────────────────────────

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_SUBTOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
_TOTAL_FILL = PatternFill("solid", fgColor="BDD7EE")


def _style_cell(cell, bold=False, fill=None, align="center"):
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _BORDER
    if fill:
        cell.fill = fill


def _num_fmt(cell):
    """Apply Romanian number format (2 decimal places)."""
    cell.number_format = "#,##0.00"


# ── Column definitions ─────────────────────────────────────────────────────

_COLUMNS = [
    ("Marca", 8),
    ("Ø [mm]", 8),
    ("Oțel", 9),
    ("Buc.", 7),
    ("Lung. [m]", 10),
    ("Lung./Ø [m]", 12),
    ("Masa Ø/m [kg/m]", 15),
    ("Masa/Ø [kg]", 12),
    ("Masa totală [kg]", 16),
]


def export_schedule(
    rows: list[dict],
    output_path: Path,
    project_name: str = "",
    project_number: str = "",
    beneficiary: str = "",
    location: str = "",
) -> None:
    """Write *rows* to an .xlsx file at *output_path*."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Extras armătură"

    current_row = 1

    # ── Project info header ───────────────────────────────────────────
    if any([project_name, project_number, beneficiary, location]):
        info_lines = [
            ("Proiect:", project_name or "—"),
            ("Nr. proiect:", project_number or "—"),
            ("Beneficiar:", beneficiary or "—"),
            ("Locație:", location or "—"),
        ]
        for label, value in info_lines:
            ws.cell(current_row, 1, label).font = Font(bold=True, size=10)
            ws.cell(current_row, 2, value).font = Font(size=10)
            current_row += 1
        current_row += 1  # blank row

    # ── Column headers ────────────────────────────────────────────────
    header_row = current_row
    for col_idx, (title, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(header_row, col_idx, title)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 30
    current_row += 1

    # ── Group rows by diameter ────────────────────────────────────────
    by_diameter: dict[int, list[dict]] = defaultdict(list)
    for row in rows:
        by_diameter[row["diameter"]].append(row)

    grand_total_weight = 0.0

    for diameter in sorted(by_diameter.keys()):
        dia_rows = by_diameter[diameter]
        dia_total_length = 0.0
        dia_total_weight = 0.0

        for r in dia_rows:
            mark_val = r.get("mark") or ""
            total_length = r.get("total_length", 0.0)
            weight = r.get("weight", 0.0)
            wpm = r.get("weight_per_meter", 0.0)
            length = r.get("length", 0.0)
            count = r.get("count", 0)

            dia_total_length += total_length
            dia_total_weight += weight

            values = [
                mark_val,
                diameter,
                r.get("steel_type", "BST500"),
                count,
                length,
                total_length,
                wpm,
                weight,
                "",  # Masa totală — only on subtotal / grand total
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(current_row, col_idx, val)
                _style_cell(cell, align="center")
                if isinstance(val, float):
                    _num_fmt(cell)
            current_row += 1

        # Subtotal row per diameter
        dia_total_weight = round(dia_total_weight, 3)
        dia_total_length = round(dia_total_length, 3)
        grand_total_weight += dia_total_weight

        subtotal_values = [
            f"Total Ø{diameter}",
            diameter,
            "",
            "",
            "",
            dia_total_length,
            "",
            dia_total_weight,
            "",
        ]
        for col_idx, val in enumerate(subtotal_values, start=1):
            cell = ws.cell(current_row, col_idx, val)
            _style_cell(cell, bold=True, fill=_SUBTOTAL_FILL, align="center")
            if isinstance(val, float):
                _num_fmt(cell)
        current_row += 1

    # ── Grand total row ───────────────────────────────────────────────
    grand_total_weight = round(grand_total_weight, 3)
    total_values = ["TOTAL", "", "", "", "", "", "", "", grand_total_weight]
    for col_idx, val in enumerate(total_values, start=1):
        cell = ws.cell(current_row, col_idx, val)
        _style_cell(cell, bold=True, fill=_TOTAL_FILL, align="center")
        if isinstance(val, float):
            _num_fmt(cell)

    # Freeze header row
    ws.freeze_panes = ws.cell(header_row + 1, 1)

    wb.save(str(output_path))
    logger.info("Exported schedule (%d rows) to %s", len(rows), output_path)
